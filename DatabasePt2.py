
'''
New Project:

Prompt user to input the # of elements they want (at least 5)

Ask them to input the rare elements they are looking for (at least 1)


(x x x x _) Al5 O12

Find the combinations of those rare elements and concantene the material at the end
'''
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

class Search:
    file = '/Users/aisling/downloads/SMaRT/Ionic_Average_Project.xlsx'
    df1 = pd.read_excel(file, sheet_name= 'List of Elements')
    df2 = pd.read_excel(file,sheet_name= 'Ionic Average Project')

    wb = openpyxl.load_workbook(file)
    ws = wb['Ionic Average Project']

    fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    def __init__(self, size_input, re_elements):
        self.size_input = size_input
        self.re_elements = re_elements.split()

    def validation(self):
        if self.size_input < 5:
            print("Size must be at least 5.")
            exit()
        column_values = self.df1['Symbol'].dropna().unique()
        nonexistent = [elem for elem in self.re_elements if elem not in column_values]

        if nonexistent:
            print("One or more of the elements do not exist.")
            return False
        else:
            return True

    def check_row(self, row, elements, threshold=4):
        columns = row.iloc[:self.size_input]
        count = sum(1 for elem in elements if elem in columns.values)
        return count >= threshold

    def find_combination(self, size, elements):
        for _, row in self.df2.iterrows():
            row_elements = row[:5].dropna().tolist()
            if all(elem in row_elements for elem in elements):
                if len(row_elements) == size:
                    return row['Concatenate'], row['Average Ionic Radius (CN = 8, Å)']
                elif len(row_elements) > size:
                    additional_elements = [elem for elem in row_elements if elem not in elements][:size - len(elements)]
                    return additional_elements, row['Average Ionic Radius (CN = 8, Å)']
        return None, None

if __name__ == "__main__":
    size_input = int(input("\nEnter the size of elements you wish for: "))
    re_input = input("Enter the rare earth elements you look for: ")
    search_instance = Search(size_input, re_input)

    if search_instance.validation():
        additional_elements, radius = search_instance.find_combination(size_input, search_instance.re_elements)
        if additional_elements:
            print(f"Complete compound: {search_instance.re_elements + list(additional_elements)}")
            print(f"Average Ionic Radius: {radius} Å")
        else:
            print("No matching compound found.")
