'''
Program that highlights elements inputted by user from the database
usecase:
Sc Y La Ce Nd 
Sc La Ce Nd Y

Important things to note:
 - order does not matter
 - user input must contain at least 5 elements
 - output will be the combinations that contain the inputted elements & the average ionic radius of those highlighted
'''

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

file = '/Users/aisling/downloads/SMaRT/Ionic_Average_Project.xlsx'
df1 = pd.read_excel(file, sheet_name= 'List of Elements')
df2 = pd.read_excel(file,sheet_name= 'Ionic Average Project')

wb = openpyxl.load_workbook(file)
ws = wb['Ionic Average Project']

fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

user_input = input("\nEnter at least 4 elements: ")
print("\n")
user_elements = user_input.split()


# checks if elements exist
def validation(user_elements):
    column_values = df1['Symbol'].dropna().unique()
    nonexistent = [elem for elem in user_elements if elem not in column_values]

    if nonexistent:
        print("The following elements do not exist.")
        for elem in nonexistent:
            print(elem)
            exit()
    else:
        return user_elements

#check if at least n elements are present in the row
def check_row(row, elements, threshold=4):
    validation(user_elements)
    columns = row.iloc[:5]
    count = sum(1 for elem in elements if elem in columns.values)
    return count >= threshold

# print rows that match the criteria
print("The following rows contain the elements you are looking for.\n")
matching_rows = df2.apply(lambda row: check_row(row, user_elements), axis=1)
highlighted_df = df2[matching_rows]
total_matching_rows = len(highlighted_df)
print(highlighted_df)
print(f"Total count: {total_matching_rows}\n")


for index, row in df2[matching_rows].iterrows():
    for col_index in range(1, len(row)+1):
        ws.cell(row=index + 2, column=col_index).fill = fill

print("Data with highlighted rows that contain at least 3 of those rows is located at highlighted_output.xlsx")
output_file = '/Users/aisling/downloads/SMaRT/highlighted_rows.xlsx'
wb.save(output_file)

print("Solely the elements that contain at least 3 elements are located at output.xlsx\n")
highlighted_df.to_excel('/Users/aisling/downloads/SMaRT/output.xlsx', index=False)



'''
New Project:

Prompt user to input the # of elements they want (at least 5)

Ask them to input the rare elements they are looking for (at least 1)


(x x x x _) Al5 O12

Find the combinations of those rare elements and concantene the material at the end

'''